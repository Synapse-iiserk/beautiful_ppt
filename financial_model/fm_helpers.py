"""Helper utilities for iFiNN Financial Model generation."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# Color scheme
DARK_GREEN = "1B4332"
MED_GREEN = "2D6A4F"
LIGHT_GREEN = "D8F3DC"
ACCENT_GREEN = "10B981"
DARK_BG = "0F1A12"
HEADER_FILL = PatternFill("solid", fgColor=DARK_GREEN)
SUBHEADER_FILL = PatternFill("solid", fgColor=MED_GREEN)
LIGHT_FILL = PatternFill("solid", fgColor=LIGHT_GREEN)
INPUT_FILL = PatternFill("solid", fgColor="E8F5E9")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
TITLE_FONT = Font(color=DARK_GREEN, bold=True, size=16)
BLUE_FONT = Font(color="0000FF", size=11)  # hardcoded inputs
GREEN_FONT = Font(color="008000", size=11)  # external refs
BLACK_FONT = Font(size=11)
BOLD_FONT = Font(bold=True, size=11)
PCT_FMT = '0.0%'
NUM_FMT = '#,##0'
INR_FMT = '₹#,##0'
INR_DEC = '₹#,##0.00'
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

YEARS = ["FY2026F", "FY2027F", "FY2028F"]
YEAR_COLS = [3, 4, 5]  # C, D, E columns

def setup_sheet(ws, title, col_widths=None):
    ws.sheet_properties.tabColor = ACCENT_GREEN
    ws.merge_cells('A1:E1')
    c = ws['A1']
    c.value = title
    c.font = Font(color=DARK_GREEN, bold=True, size=16)
    c.alignment = CENTER
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

def write_header_row(ws, row, headers, fill=HEADER_FILL, font=WHITE_FONT):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = font
        c.fill = fill
        c.alignment = CENTER
        c.border = THIN_BORDER

def write_row(ws, row, values, fonts=None, fmts=None, bold_first=True):
    for j, v in enumerate(values, 1):
        c = ws.cell(row=row, column=j, value=v)
        c.border = THIN_BORDER
        c.alignment = LEFT if j == 1 else CENTER
        if bold_first and j == 1:
            c.font = BOLD_FONT
        elif fonts and j <= len(fonts) and fonts[j-1]:
            c.font = fonts[j-1]
        if fmts and j <= len(fmts) and fmts[j-1]:
            c.number_format = fmts[j-1]

def write_section(ws, start_row, section_title, headers, data_rows, fmt=None):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
    c = ws.cell(row=start_row, column=1, value=section_title)
    c.font = WHITE_FONT
    c.fill = SUBHEADER_FILL
    c.alignment = LEFT
    for col in range(1, len(headers)+1):
        ws.cell(row=start_row, column=col).fill = SUBHEADER_FILL
        ws.cell(row=start_row, column=col).border = THIN_BORDER
    start_row += 1
    write_header_row(ws, start_row, headers)
    start_row += 1
    for vals in data_rows:
        fmts = [None] + [fmt]*(len(vals)-1) if fmt else None
        write_row(ws, start_row, vals, fmts=fmts)
        start_row += 1
    return start_row
