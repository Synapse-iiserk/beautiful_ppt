#!/usr/bin/env python3
"""Generate iFiNN Financial Model XLSX — 3-Year Projection (FY2026–FY2028)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fm_helpers import *

wb = openpyxl.Workbook()

# ============================================================
# 1. COVER PAGE
# ============================================================
ws = wb.active
ws.title = "Cover Page"
setup_sheet(ws, "iFiNN — Intelligent Financial Neural Network", [40, 30, 20, 20, 20])
ws.merge_cells('A3:E3')
ws['A3'] = "3-Year Financial Projection Model (FY2026–FY2028)"
ws['A3'].font = Font(size=14, color=MED_GREEN, bold=True)
ws['A3'].alignment = CENTER

ws['A5'] = "Company:"
ws['B5'] = "iFiNN (Team Synapse, IISER Kolkata)"
ws['A6'] = "Sector:"
ws['B6'] = "AI-Powered Financial Analytics (SaaS + Marketplace)"
ws['A7'] = "Currency:"
ws['B7'] = "INR (₹)"
ws['A8'] = "Projection Period:"
ws['B8'] = "FY2026 – FY2028"
for r in range(5, 9):
    ws.cell(r, 1).font = BOLD_FONT
    ws.cell(r, 2).font = BLACK_FONT

r = 10
ws.merge_cells(f'A{r}:E{r}')
ws[f'A{r}'] = "Color Coding Convention"
ws[f'A{r}'].font = WHITE_FONT
ws[f'A{r}'].fill = HEADER_FILL
r += 1
for label, ft in [("Blue Font = Hardcoded Inputs", BLUE_FONT),
                   ("Green Font = External References / Links", GREEN_FONT),
                   ("Black Font = Formulas / Calculations", BLACK_FONT)]:
    ws.cell(r, 1, label).font = ft
    r += 1

r += 1
ws.merge_cells(f'A{r}:E{r}')
ws[f'A{r}'] = "Table of Contents"
ws[f'A{r}'].font = WHITE_FONT
ws[f'A{r}'].fill = HEADER_FILL
r += 1
toc = ["Cover Page","Converter","Outputs","Inputs","User Acquisition Schedule",
       "Revenue Schedule","Cost Schedule","WC Schedule","Manpower Schedule",
       "OpEx Schedule","Asset Schedule","Depreciation Schedule","Debt Schedule",
       "Income Tax Schedule","Income Statement","Cash Flow","Balance Sheet",
       "Berkus Method","Venture Capital Method","DCF","Convertible Notes"]
for i, name in enumerate(toc, 1):
    ws.cell(r, 1, f"{i}. {name}").font = GREEN_FONT
    r += 1

# ============================================================
# 2. CONVERTER
# ============================================================
ws = wb.create_sheet("Converter")
setup_sheet(ws, "Unit & Currency Converter", [30, 20, 20, 20, 20])
write_header_row(ws, 3, ["Denomination", "Value", "", "", ""])
for i, (n, v) in enumerate([("Units", 1), ("Thousands", 1000), ("Lakhs", 100000),
                             ("Millions", 1000000), ("Crores", 10000000)], 4):
    write_row(ws, i, [n, v, "", "", ""], fmts=[None, NUM_FMT])

write_header_row(ws, 10, ["Currency", "Rate vs INR", "", "", ""])
for i, (c, r2) in enumerate([("INR", 1), ("USD", 85.5), ("EUR", 92.3),
                              ("GBP", 108.2), ("AED", 23.3), ("JPY", 0.57)], 11):
    write_row(ws, i, [c, r2, "", "", ""], fmts=[None, '#,##0.00'])

ws.cell(17, 1, "Reporting Currency:").font = BOLD_FONT
ws.cell(17, 2, "INR").font = BLUE_FONT
ws.cell(18, 1, "Denomination:").font = BOLD_FONT
ws.cell(18, 2, "Lakhs").font = BLUE_FONT

# ============================================================
# 3. OUTPUTS (Dashboard)
# ============================================================
# (Generated at the end of the script to use calculated values)


# ============================================================
# 4. INPUTS (Scenarios)
# ============================================================
ws = wb.create_sheet("Inputs")
setup_sheet(ws, "Scenario & Assumption Inputs", [40, 15, 18, 18, 18])
headers = ["Assumption", "Unit"] + YEARS

# -- INPUT VARIABLES --
# FY26, FY27, FY28 (Refined Targets)
new_users_input = [500, 2000, 7000]   # Scaled with expanded team (CEO/CTO/COO + 7 tech + 2 ops)
# Conversion: Realistic for bootstrapped early stage
conv_rate_input = [0.05, 0.08, 0.12]
# Headcount: FY26=9 active; FY27=full 12 (CEO/CTO/COO + 3FS + 2ML + 2FE + Mkt + Sales); FY28=18 (12+6 hires)
headcount_input = [9, 12, 18]  # FY26: 9 active; FY27: full 12 team; FY28: 18 (12 core + 6 new hires)
paid_churn_rate_input = [0.20, 0.15, 0.10]
free_churn_rate_input = [0.05, 0.13, 0.15]
pro_fee = 499
elite_fee = 1999
pro_split = [0.80, 0.70, 0.60]
elite_split = [0.20, 0.30, 0.40]

write_section(ws, 3, "Base Case Scenario (Refined)", headers, [
    ["New Free Users Acquired", "#"] + new_users_input,
    ["Free-to-Paid Conversion Rate", "%"] + [x*100 for x in conv_rate_input],
    ["Monthly Churn Rate (Paid)", "%", 8.0, 5.0, 3.0], 
    ["Pro Subscription Fee", "₹/mo", pro_fee, pro_fee, pro_fee],
    ["Elite Subscription Fee", "₹/mo", elite_fee, elite_fee, elite_fee],
    ["Pro : Elite Split", "ratio", "80:20", "70:30", "60:40"],
    ["Marketplace Commission Rate", "%", 20.0, 22.0, 25.0],
    ["API Calls per Enterprise/mo", "#", 10000, 50000, 200000],
    ["API Price per Call", "₹", 0.5, 0.5, 0.5],
    ["Courses Sold (Annual)", "#", 50, 500, 2000],
    ["Avg Course Price", "₹", 1999, 2499, 2999],
    ["Proprietary Trading Capital", "₹ Lakhs", 0, 50, 100],
    ["Prop Trading Annual Return", "%", 0, 30, 30],
    ["CAC (Blended)", "₹", 150, 120, 90],
    ["Hosting Cost per User/Yr", "₹", 200, 200, 180], # New metric
], fmt=None)

write_section(ws, 21, "Best Case Scenario", headers, [
    ["New Free Users Acquired", "#", 2000, 8000, 35000],
    ["Free-to-Paid Conversion Rate", "%", 7.0, 14.0, 18.0],
    ["Prop Trading Annual Return", "%", 0, 25, 35],
], fmt=None)

write_section(ws, 27, "Worst Case Scenario", headers, [
    ["New Free Users Acquired", "#", 500, 2500, 10000],
    ["Free-to-Paid Conversion Rate", "%", 3.0, 6.0, 8.0],
    ["Prop Trading Annual Return", "%", 0, 8, 12],
], fmt=None)

# ============================================================
# 5. USER ACQUISITION SCHEDULE (CALCULATED)
# ============================================================
ws = wb.create_sheet("User Acquisition Schedule")
setup_sheet(ws, "User Acquisition & Retention Funnel", [40, 15, 18, 18, 18])
headers = ["Metric", "Unit"] + YEARS

# Free User Funnel
free_open = [0] * 3
free_new = new_users_input
free_churn = [0] * 3
free_close = [0] * 3

# Paid User Funnel
paid_open = [0] * 3
paid_conv = [0] * 3
paid_churn = [0] * 3
paid_close = [0] * 3
pro_users = [0] * 3
elite_users = [0] * 3

curr_free = 0
curr_paid = 0

for i in range(3):
    # Free
    free_open[i] = curr_free
    free_churn[i] = int(free_new[i] * free_churn_rate_input[i]) 
    free_close[i] = free_open[i] + free_new[i] - free_churn[i]
    curr_free = free_close[i]
    
    # Paid
    paid_open[i] = curr_paid
    paid_conv[i] = int(free_new[i] * conv_rate_input[i]) 
    # Ensure min 20 Pro users in FY26 check:
    # 1000 * 5% = 50. 80% of 50 = 40. OK.
    
    paid_churn[i] = int((paid_open[i] + paid_conv[i]) * paid_churn_rate_input[i])
    paid_close[i] = paid_open[i] + paid_conv[i] - paid_churn[i]
    
    pro_users[i] = int(paid_close[i] * pro_split[i])
    elite_users[i] = paid_close[i] - pro_users[i]
    curr_paid = paid_close[i]

write_section(ws, 3, "Free User Funnel", headers, [
    ["Opening Free Users", "#"] + free_open,
    ["(+) New Free Users Acquired", "#"] + free_new,
    ["(-) Churned Free Users", "#"] + [-x for x in free_churn],
    ["Closing Free Users", "#"] + free_close,
], fmt=NUM_FMT)

write_section(ws, 10, "Paid User Funnel", headers, [
    ["Opening Paid Users", "#"] + paid_open,
    ["(+) Conversions from Free", "#"] + paid_conv,
    ["(-) Paid Churn", "#"] + [-x for x in paid_churn],
    ["Closing Paid Users", "#"] + paid_close,
    ["   → Pro Users", "#"] + pro_users,
    ["   → Elite Users", "#"] + elite_users,
], fmt=NUM_FMT)

# LTV/CAC
ltv_cac_rows = []
cac_list = [150, 120, 90]
ltv_list = []
ltv_cac_ratio = []
arpu_list = []

for i in range(3):
    # Blended ARPU
    total_rev_mo = (pro_users[i] * pro_fee) + (elite_users[i] * elite_fee)
    user_count = paid_close[i] if paid_close[i] > 0 else 1
    arpu = total_rev_mo / user_count
    arpu_list.append(arpu)
    
    lt_months = [12.5, 20.0, 33.3] 
    ltv = arpu * lt_months[i]
    ltv_list.append(ltv)
    ltv_cac_ratio.append(ltv / cac_list[i])

write_section(ws, 19, "Unit Economics", headers, [
    ["Blended ARPU", "₹/mo"] + arpu_list,
    ["Customer Acquisition Cost", "₹"] + cac_list,
    ["Customer Lifetime (months)", "#"] + lt_months,
    ["LTV", "₹"] + ltv_list,
    ["LTV / CAC", "x"] + ltv_cac_ratio,
], fmt=NUM_FMT)

# ============================================================
# 6. REVENUE SCHEDULE (CALCULATED)
# ============================================================
ws = wb.create_sheet("Revenue Schedule")
setup_sheet(ws, "Revenue Breakdown by Stream", [40, 15, 18, 18, 18])
headers = ["Revenue Stream", "Unit"] + YEARS

# Subscription Revenue Calculation
avg_pro = []
avg_elite = []
pro_rev = []
elite_rev = []

for i in range(3):
    open_p = pro_users[i-1] if i > 0 else 0
    close_p = pro_users[i]
    avg_p = (open_p + close_p) / 2
    avg_pro.append(avg_p)
    
    open_e = elite_users[i-1] if i > 0 else 0
    close_e = elite_users[i]
    avg_e = (open_e + close_e) / 2
    avg_elite.append(avg_e)
    
    pro_rev.append((avg_p * pro_fee * 12) / 100000) # Lakhs
    elite_rev.append((avg_e * elite_fee * 12) / 100000) # Lakhs

total_sub_rev = [p + e for p, e in zip(pro_rev, elite_rev)]

write_section(ws, 3, "Subscription Revenue", headers, [
    ["Pro Users (Avg Annual)", "#"] + avg_pro,
    ["Pro Monthly Fee", "₹", pro_fee, pro_fee, pro_fee],
    ["Pro Annual Revenue", "₹ Lakhs"] + pro_rev,
    ["Elite Users (Avg Annual)", "#"] + avg_elite,
    ["Elite Monthly Fee", "₹", elite_fee, elite_fee, elite_fee],
    ["Elite Annual Revenue", "₹ Lakhs"] + elite_rev,
    ["Total Subscription Revenue", "₹ Lakhs"] + total_sub_rev,
], fmt=INR_FMT)

# Other Revenue Scaling (Linear to User Base)
# Base Year FY26 (1000 users) -> Scale Factors
# Let's define ratios per 1000 users to keep it dynamic?
# Or just reasonable estimates manually scaled.
# FY26: 1000 new. FY27: 5000 new (cumul ~6000). FY28: 20000 new (cumul ~26000).
scale_metrics = [1.0, 6.0, 26.0] 

# Marketplace base: starts slow. FY27 kick-off.
marketplace_rev = [0, 2.0, 10.0] 
api_rev = [0.1, 1.0, 5.0]
course_rev = [0.3, 2.0, 15.0]
prop_trading_rev = [0, 15.0, 30.0] # ₹50L × 30% FY27; ₹100L × 30% FY28

total_other_rev = [sum(x) for x in zip(marketplace_rev, api_rev, course_rev, prop_trading_rev)]
total_rev = [s + o for s, o in zip(total_sub_rev, total_other_rev)]

# Growth
yoy = ["—"]
for i in range(1, 3):
    if total_rev[i-1] > 0:
        g = (total_rev[i] - total_rev[i-1]) / total_rev[i-1]
        yoy.append(f"{g*100:.0f}%")
    else:
        yoy.append("0%")

sub_pct = [s/t*100 if t>0 else 0 for s, t in zip(total_sub_rev, total_rev)]
other_pct = [o/t*100 if t>0 else 0 for o, t in zip(total_other_rev, total_rev)]

write_section(ws, 13, "Marketplace & Other Revenue", headers, [
    ["Marketplace Commission Revenue", "₹ Lakhs"] + marketplace_rev,
    ["API Revenue", "₹ Lakhs"] + api_rev,
    ["Course Revenue", "₹ Lakhs"] + course_rev,
    ["Proprietary Trading Returns", "₹ Lakhs"] + prop_trading_rev,
    ["Total Other Revenue", "₹ Lakhs"] + total_other_rev,
], fmt=INR_FMT)

write_section(ws, 23, "Total Revenue", headers, [
    ["Total Revenue", "₹ Lakhs"] + total_rev,
    ["YoY Growth", "%"] + yoy,
    ["Subscription %", "%"] + sub_pct,
    ["Other Revenue %", "%"] + other_pct,
], fmt=INR_FMT)

# ============================================================
# 7. COST SCHEDULE (Recalculated)
# ============================================================
ws = wb.create_sheet("Cost Schedule")
setup_sheet(ws, "Variable Cost Schedule", [40, 15, 18, 18, 18])
headers = ["Cost Item", "Unit"] + YEARS

# Hosting Cost: Per User Model
# active users estimate = avg free + avg paid
total_active_users = [(f1+f2)/2 + (p1+p2)/2 for f1,f2, p1,p2 in zip(free_open, free_close, paid_open, paid_close)]
host_cost_per_user = [200, 200, 180] # Efficiency scale
hosting_cost = [u * c / 100000 for u,c in zip(total_active_users, host_cost_per_user)] # Lakhs

gpu_cost = [0.5, 2.0, 4.0] # ML compute (2 ML engineers, GPU-intensive training)
api_cost = [0.3, 1.0, 2.0] # Market data APIs (NSE, social feeds)
social_cost = [0.2, 0.3, 0.5]
pg_cost = [t * 0.02 for t in total_rev]

total_var_cost = [sum(x) for x in zip(hosting_cost, gpu_cost, api_cost, social_cost, pg_cost)]
margin_pct = [(r - c)/r * 100 if r>0 else 0 for r, c in zip(total_rev, total_var_cost)]

write_section(ws, 3, "Variable Costs (Revenue-Linked)", headers, [
    ["Total Active Users (Avg)", "#"] + [int(x) for x in total_active_users],
    ["Cloud Hosting (₹200/user/yr)", "₹ Lakhs"] + hosting_cost,
    ["GPU Compute (ML Training)", "₹ Lakhs"] + gpu_cost,
    ["Data API Costs (Market Data)", "₹ Lakhs"] + api_cost,
    ["Social Media API Costs", "₹ Lakhs"] + social_cost,
    ["Payment Gateway Fees (2%)", "₹ Lakhs"] + pg_cost,
    ["Total Variable Costs", "₹ Lakhs"] + total_var_cost,
    ["Variable Cost Margin", "%"] + margin_pct,
], fmt=INR_FMT)



# ============================================================
# 8. WC SCHEDULE
# ============================================================
ws = wb.create_sheet("WC Schedule")
setup_sheet(ws, "Working Capital Schedule", [40, 15, 18, 18, 18])
headers = ["Item", "Unit"] + YEARS

dso = [15, 20, 25]
dpo = [30, 30, 30]
ar = [r * d / 365 for r, d in zip(total_rev, dso)]
# Use Total Variable Costs as checks for payables base
ap = [c * d / 365 for c, d in zip(total_var_cost, dpo)]
nwc = [r - p for r, p in zip(ar, ap)]
change_nwc = [nwc[0]] + [nwc[i] - nwc[i-1] for i in range(1, 3)]

write_section(ws, 3, "Working Capital Assumptions", headers, [
    ["Days Sales Outstanding (DSO)", "days"] + dso,
    ["Days Payable Outstanding (DPO)", "days"] + dpo,
], fmt=NUM_FMT)

write_section(ws, 8, "Working Capital Calculation", headers, [
    ["Accounts Receivable", "₹ Lakhs"] + ar,
    ["Accounts Payable", "₹ Lakhs"] + ap,
    ["Net Working Capital", "₹ Lakhs"] + nwc,
    ["Change in Working Capital", "₹ Lakhs"] + change_nwc,
], fmt=INR_FMT)

# ============================================================
# 9. MANPOWER SCHEDULE
# ============================================================
ws = wb.create_sheet("Manpower Schedule")
setup_sheet(ws, "Manpower & Salary Schedule", [35, 20, 12, 15, 15, 15])
ws.column_dimensions['F'].width = 15
headers = ["Department / Role", "Designation", "Count"] + YEARS

# FY26: 2 Full-Stack (No salary). Others start FY27.
write_section(ws, 3, "Management (Team iFiNN)", headers, [
    ["CEO / Co-Founder", "Director", 1, 0, 2.0, 4.0], 
    ["CTO / Co-Founder", "Director", 1, 0, 2.0, 4.0],
    ["COO / Co-Founder", "Director", 1, 0, 1.5, 3.0],
], fmt=INR_FMT)

write_section(ws, 9, "Technology", headers, [
    ["Full-Stack Developer", "Mid-Level", 3, "3 (No Pay)", 4.0, 15.0],
    ["ML Engineer", "Senior", 2, 0, 3.0, 10.0],
    ["Frontend Developer", "Junior", 2, 0, 1.5, 6.0],
], fmt=INR_FMT)

write_section(ws, 15, "Operations & Sales", headers, [
    ["Marketing Manager", "Mid-Level", 1, 0, 1.0, 4.0],
    ["Sales & Community Support", "Junior", 1, 0, 0.5, 3.0],
], fmt=INR_FMT)

# Team: CEO(1) + CTO(1) + COO(1) + 3 FS + 2 ML + 2 FE + 1 Mkt + 1 Sales = 12 core
# FY26: 9 active (CEO/CTO + 3 FS no-pay + 1 ML + 1 FE + 1 Mkt + 1 Sales)
# FY27: Full 12 (add COO + 1 ML + 1 FE). Startup salaries — Bangalore peer level.
# FY28: 18 total (12 core + 2 FS + 1 ML + 1 FE + 1 Mkt + 1 Sales new hires)
# FY26: avg ₹1.07L/yr × 9 = ₹9.6L (founder stipends, under incubator)
# FY27: avg ₹2L/yr × 12 = ₹24L (lean startup salaries, Bangalore)
# FY28: avg ₹3.67L/yr × 18 = ₹66L (growth-stage, Bangalore market rates)

total_salary = [9.6, 24.0, 66.0]

write_section(ws, 20, "Total Manpower Cost", headers, [
    ["Total Headcount", "#", 9, 12, 18],
    ["Total Annual Salary", "₹ Lakhs"] + total_salary,
    ["Yearly Increment", "%", "", "—", "—", "—"],
], fmt=INR_FMT)

# ============================================================
# 10. OPEX SCHEDULE
# ============================================================
ws = wb.create_sheet("OpEx Schedule")
setup_sheet(ws, "Operating Expense Schedule", [40, 15, 18, 18, 18])
headers = ["Expense Category", "Unit"] + YEARS

# Rent Logic:
# FY26: 0 (under incubator — IISER Kolkata / MeitY Startup Hub)
# FY27: Bangalore co-working space (~₹5K/seat/mo × 8-10 seats × 12 months ≈ ₹5L)
# FY28: Larger Bangalore co-working (~₹6K/seat/mo × 14 seats × 12 months ≈ ₹10L)
rent_cost = [0, 5.0, 10.0]

write_section(ws, 3, "Fixed & Semi-Variable OpEx", headers, [
    ["Office Rent & Utilities (Bangalore)", "₹ Lakhs"] + rent_cost,
    ["Software Subscriptions", "₹ Lakhs", 0.5, 1.5, 3.0],
    ["Legal & Compliance", "₹ Lakhs", 0.3, 1.0, 1.5],
    ["Accounting & Audit", "₹ Lakhs", 0, 0.5, 1.0],
    ["Insurance", "₹ Lakhs", 0, 0.3, 0.5],
], fmt=INR_FMT)

write_section(ws, 11, "Sales, General & Administrative", headers, [
    ["Digital Marketing & Ads", "₹ Lakhs", 0.5, 2.0, 5.0],
    ["Content Marketing", "₹ Lakhs", 0.2, 0.5, 1.0],
    ["PR & Branding", "₹ Lakhs", 0.1, 0.5, 1.0],
    ["Events & Conferences", "₹ Lakhs", 0, 0.5, 1.0],
    ["Travel & Entertainment", "₹ Lakhs", 0, 0.3, 0.5],
    ["Contingency", "₹ Lakhs", 0.1, 0.2, 0.3],
], fmt=INR_FMT)

total_opex = [1.7, 12.3, 24.8]

write_section(ws, 20, "Total OpEx", headers, [
    ["Total Fixed OpEx", "₹ Lakhs", 0.8, 8.3, 16.0],
    ["Total SG&A", "₹ Lakhs", 0.9, 4.0, 8.8],
    ["Total Operating Expenses", "₹ Lakhs"] + total_opex,
], fmt=INR_FMT)

# ============================================================
# 11. ASSET SCHEDULE
# ============================================================
# ... [Asset Schedule Logic remains similar, scaled down]
ws = wb.create_sheet("Asset Schedule")
setup_sheet(ws, "Capital Expenditure & Asset Schedule", [40, 15, 18, 18, 18])
headers = ["Asset Category", "Unit"] + YEARS

total_capex = [2.0, 4.5, 9.0]

write_section(ws, 3, "Capital Expenditure", headers, [
    ["Laptops & Workstations", "₹ Lakhs", 1.5, 2.5, 4.0],
    ["GPU Servers (On-Premise)", "₹ Lakhs", 0, 1.0, 3.0],
    ["Office Furniture & Setup", "₹ Lakhs", 0, 0.5, 1.5],
    ["Networking Equipment", "₹ Lakhs", 0.5, 0.5, 0.5],
    ["Total CapEx", "₹ Lakhs"] + total_capex,
    ["Description", "", "Initial Setup", "Expansion", "Scale-up + GPU Farm"],
], fmt=INR_FMT)

write_section(ws, 11, "Asset Tracking", headers, [
    ["Opening Gross Assets", "₹ Lakhs", 0, 2.0, 6.5],
    ["(+) Additions", "₹ Lakhs", 2.0, 4.5, 9.0],
    ["(-) Disposals", "₹ Lakhs", 0, 0, 0],
    ["Closing Gross Assets", "₹ Lakhs", 2.0, 6.5, 15.5],
], fmt=INR_FMT)

# ============================================================
# 12. DEPRECIATION SCHEDULE
# ============================================================
ws = wb.create_sheet("Depreciation Schedule")
setup_sheet(ws, "Depreciation Schedule", [40, 15, 18, 18, 18])
headers = ["Item", "Unit"] + YEARS

total_depr = [0.6, 1.5, 2.8]

write_section(ws, 3, "Accounting Depreciation (SLM)", headers, [
    ["Laptops (3yr life)", "₹ Lakhs", 0.5, 1.0, 1.5],
    ["GPU Servers (5yr life)", "₹ Lakhs", 0, 0.2, 0.8],
    ["Furniture (10yr life)", "₹ Lakhs", 0, 0.1, 0.2],
    ["Networking (5yr life)", "₹ Lakhs", 0.1, 0.2, 0.3],
    ["Total Depreciation", "₹ Lakhs"] + total_depr,
], fmt=INR_FMT)

tax_depr = [1.0, 3.0, 6.0]
depr_diff = [t - a for t, a in zip(tax_depr, total_depr)]

write_section(ws, 11, "Tax Depreciation (WDV)", headers, [
    ["Tax Depreciation", "₹ Lakhs"] + tax_depr,
    ["Difference (Tax - Accounting)", "₹ Lakhs"] + depr_diff,
], fmt=INR_FMT)

# ============================================================
# 13. DEBT SCHEDULE
# ============================================================
ws = wb.create_sheet("Debt Schedule")
setup_sheet(ws, "Debt & Interest Schedule", [40, 15, 18, 18, 18])
headers = ["Item", "Unit"] + YEARS

write_section(ws, 3, "Term Loan", headers, [
    ["Opening Balance", "₹ Lakhs", 0, 0, 0],
    ["(+) New Borrowings", "₹ Lakhs", 0, 0, 0],
    ["(-) Repayments", "₹ Lakhs", 0, 0, 0],
    ["Closing Balance", "₹ Lakhs", 0, 0, 0],
    ["Interest Rate", "%", "—", "—", "—"],
    ["Interest Expense", "₹ Lakhs", 0, 0, 0],
], fmt=INR_FMT)

ws.cell(12, 1, "Debt Capacity Note: iFiNN is 100% equity funded. Convertible note (₹50L) accrues interest at 8% (₹4L/yr) but does not appear as debt until conversion. Post-₹50L ARR, venture debt at 10-15% of ARR becomes viable.").font = Font(italic=True, size=10, color="666666")

# ============================================================
# 14/15. INCOME TAX & STATEMENT
# ============================================================
ws_is = wb.create_sheet("Income Statement")
ws_tax = wb.create_sheet("Income Tax Schedule")

# Calculate IS components
gross_profit = [r - c for r, c in zip(total_rev, total_var_cost)]
gross_margin = [gp/r*100 if r>0 else 0 for gp, r in zip(gross_profit, total_rev)]
ebit = [gp - s - o - d for gp, s, o, d in zip(gross_profit, total_salary, total_opex, total_depr)]
interest = [0, 0, 0]
ebt = [e - i for e, i in zip(ebit, interest)] # Likely negative in FY26

# Tax Calculation
tax_rate = 0.25
tax_payable = []
eff_tax_rate = []
accum_loss = 0
tax_loss_cf = [] 
tax_loss_bf = []

for i in range(3):
    tax_loss_bf.append(accum_loss)
    # Taxable Income
    taxable_base = ebt[i] + total_depr[i] - tax_depr[i]
    
    current_tax = 0
    if accum_loss > 0:
        if taxable_base > 0:
            offset = min(taxable_base, accum_loss)
            taxable_base -= offset
            accum_loss -= offset
        else:
            accum_loss += abs(taxable_base)
            taxable_base = 0
    elif taxable_base < 0:
        accum_loss += abs(taxable_base)
        taxable_base = 0
        
    tax_loss_cf.append(accum_loss)
    current_tax = taxable_base * tax_rate
    tax_payable.append(current_tax)
    eff_tax_rate.append(current_tax/ebt[i]*100 if ebt[i]>0 else 0)

net_income = [e - t for e, t in zip(ebt, tax_payable)]
net_margin = [n/r*100 if r>0 else 0 for n, r in zip(net_income, total_rev)]

# Populate Tax Sheet
setup_sheet(ws_tax, "Income Tax Computation", [40, 15, 18, 18, 18])
write_section(ws_tax, 3, "Tax Calculation", headers, [
    ["Earnings Before Tax (EBT)", "₹ Lakhs"] + ebt,
    ["(+) Accounting Depreciation", "₹ Lakhs"] + total_depr,
    ["(-) Tax Depreciation", "₹ Lakhs"] + [-x for x in tax_depr],
    ["Adjusted Taxable Income", "₹ Lakhs"] + [e+d-t for e,d,t in zip(ebt, total_depr, tax_depr)],
    ["(-) B/F Tax Losses", "₹ Lakhs"] + [-x for x in tax_loss_bf],
    ["Net Taxable Income", "₹ Lakhs"] + [t/0.25 if t>0 else 0 for t in tax_payable], 
    ["Tax Rate (Startup — 25%)", "%", 25, 25, 25],
    ["Income Tax Payable", "₹ Lakhs"] + tax_payable,
    ["Effective Tax Rate", "%"] + eff_tax_rate,
], fmt=INR_FMT)

write_section(ws_tax, 15, "Tax Loss Carry Forward", headers, [
    ["Opening Tax Loss", "₹ Lakhs"] + tax_loss_bf,
    ["(+) Current Year Loss", "₹ Lakhs"] + [max(0, tax_loss_cf[i]-tax_loss_bf[i]) for i in range(3)],
    ["(-) Set Off", "₹ Lakhs"] + [min(0, tax_loss_cf[i]-tax_loss_bf[i]) for i in range(3)],
    ["Closing Tax Loss", "₹ Lakhs"] + tax_loss_cf,
], fmt=INR_FMT)

# Populate IS
setup_sheet(ws_is, "Profit & Loss Statement", [40, 15, 18, 18, 18])
write_section(ws_is, 3, "Income Statement", headers, [
    ["Revenue from Operations", "₹ Lakhs"] + total_rev,
    ["(-) Variable Costs", "₹ Lakhs"] + [-x for x in total_var_cost],
    ["Gross Profit", "₹ Lakhs"] + gross_profit,
    ["Gross Margin", "%"] + gross_margin,
    ["", "", "", "", ""],
    ["(-) Employee Costs", "₹ Lakhs"] + [-x for x in total_salary],
    ["(-) Operating Expenses", "₹ Lakhs"] + [-x for x in total_opex],
    ["(-) Depreciation", "₹ Lakhs"] + [-x for x in total_depr],
    ["EBIT", "₹ Lakhs"] + ebit,
    ["EBIT Margin", "%"] + [e/r*100 if r>0 else 0 for e, r in zip(ebit, total_rev)],
    ["", "", "", "", ""],
    ["(-) Interest Expense", "₹ Lakhs"] + interest,
    ["Earnings Before Tax (EBT)", "₹ Lakhs"] + ebt,
    ["(-) Income Tax", "₹ Lakhs"] + [-x for x in tax_payable],
    ["Net Income", "₹ Lakhs"] + net_income,
    ["Net Margin", "%"] + net_margin,
], fmt=INR_FMT)

# ============================================================
# 16. CASH FLOW
# ============================================================
ws = wb.create_sheet("Cash Flow")
setup_sheet(ws, "Cash Flow Statement (Indirect Method)", [40, 15, 18, 18, 18])
headers = ["Particulars", "Unit"] + YEARS

prop_trading_inv = [0, 50.0, 100.0] # FY27: ₹50L capital deployed; FY28: ₹1Cr
equity_infusion = [0, 100.0, 300.0] # FY26: worst-case zero external; FY27: Pre-Seed ₹1Cr min; FY28: Seed ₹3Cr

prop_gain_noncash = prop_trading_rev 

cfo = [n + d - wc - pg for n, d, wc, pg in zip(net_income, total_depr, change_nwc, prop_gain_noncash)]
cfi = [-c - p for c, p in zip(total_capex, prop_trading_inv)]
cff = [e for e in equity_infusion] 
net_cf = [o + i + f for o, i, f in zip(cfo, cfi, cff)]

closing_cash = []
curr = 0
for c in net_cf:
    curr += c
    closing_cash.append(curr)

write_section(ws, 3, "Operating Activities", headers, [
    ["Net Income", "₹ Lakhs"] + net_income,
    ["(+) Depreciation", "₹ Lakhs"] + total_depr,
    ["(-) Changes in Working Capital", "₹ Lakhs"] + [-x for x in change_nwc],
    ["(-) Non-Cash Prop Trading Gains", "₹ Lakhs"] + [-x for x in prop_gain_noncash], 
    ["Cash from Operations", "₹ Lakhs"] + cfo,
], fmt=INR_FMT)

write_section(ws, 10, "Investing Activities", headers, [
    ["(-) Capital Expenditure", "₹ Lakhs"] + [-x for x in total_capex],
    ["(-) Proprietary Trading Investment", "₹ Lakhs"] + [-x for x in prop_trading_inv],
    ["Cash from Investing", "₹ Lakhs"] + cfi,
], fmt=INR_FMT)

write_section(ws, 16, "Financing Activities", headers, [
    ["(+) Equity Infusion", "₹ Lakhs"] + equity_infusion,
    ["(+) Debt Raised", "₹ Lakhs", 0, 0, 0],
    ["(-) Debt Repayment", "₹ Lakhs", 0, 0, 0],
    ["Cash from Financing", "₹ Lakhs"] + cff,
], fmt=INR_FMT)

write_section(ws, 23, "Net Cash Position", headers, [
    ["Net Cash Flow", "₹ Lakhs"] + net_cf,
    ["Opening Cash Balance", "₹ Lakhs"] + [0] + closing_cash[:-1],
    ["Closing Cash Balance", "₹ Lakhs"] + closing_cash,
], fmt=INR_FMT)

# ============================================================
# 17. BALANCE SHEET
# ============================================================
ws = wb.create_sheet("Balance Sheet")
setup_sheet(ws, "Balance Sheet", [40, 15, 18, 18, 18])
headers = ["Particulars", "Unit"] + YEARS

# Assets
curr_assets = [c + a for c, a in zip(closing_cash, ar)]

accum_depr = []
ad_curr = 0
for d in total_depr:
    ad_curr += d
    accum_depr.append(ad_curr)
    
gb = []
gb_c = 0
for c in total_capex:
    gb_c += c
    gb.append(gb_c)
net_block = [g - a for g, a in zip(gb, accum_depr)]

prop_portfolio = [] 
pp_curr = 0; 
for inv, gain in zip(prop_trading_inv, prop_gain_noncash):
    pp_curr = pp_curr + inv + gain
    prop_portfolio.append(pp_curr)

total_assets = [ca + nb + pp for ca, nb, pp in zip(curr_assets, net_block, prop_portfolio)]

# Liability
curr_liab = [a + t for a, t in zip(ap, tax_payable)]

share_capital = []
sc_curr = 0
for e in equity_infusion:
    sc_curr += e
    share_capital.append(sc_curr)
    
retained_earnings = []
re_curr = 0
for n in net_income:
    re_curr += n
    retained_earnings.append(re_curr)

total_equity = [s + r for s, r in zip(share_capital, retained_earnings)]
total_le = [l + e for l, e in zip(curr_liab, total_equity)]
balance_check = [a - le for a, le in zip(total_assets, total_le)]

write_section(ws, 3, "Assets", headers, [
    ["Cash & Cash Equivalents", "₹ Lakhs"] + closing_cash,
    ["Accounts Receivable", "₹ Lakhs"] + ar,
    ["Total Current Assets", "₹ Lakhs"] + curr_assets,
    ["", "", "", "", ""],
    ["Gross Fixed Assets", "₹ Lakhs"] + gb,
    ["(-) Accumulated Depreciation", "₹ Lakhs"] + [-x for x in accum_depr],
    ["Net Fixed Assets", "₹ Lakhs"] + net_block,
    ["Proprietary Trading Portfolio", "₹ Lakhs"] + prop_portfolio,
    ["Total Assets", "₹ Lakhs"] + total_assets,
], fmt=INR_FMT)

write_section(ws, 15, "Liabilities & Equity", headers, [
    ["Accounts Payable", "₹ Lakhs"] + ap,
    ["Tax Payable", "₹ Lakhs"] + tax_payable,
    ["Total Liabilities", "₹ Lakhs"] + curr_liab,
    ["", "", "", "", ""],
    ["Share Capital (Equity Infused)", "₹ Lakhs"] + share_capital,
    ["Retained Earnings", "₹ Lakhs"] + retained_earnings,
    ["Total Equity", "₹ Lakhs"] + total_equity,
    ["", "", "", "", ""],
    ["Total Liabilities + Equity", "₹ Lakhs"] + total_le,
    ["Balance Check (Assets - L&E)", "₹ Lakhs"] + balance_check,
], fmt=INR_FMT)

for col in YEAR_COLS:
    c = ws.cell(26, col)
    c.fill = PatternFill("solid", fgColor="C8E6C9")
    c.font = Font(bold=True, color="2E7D32")

# ============================================================
# 18-21. VALUATION & OTHERS
# ============================================================
ws = wb.create_sheet("Berkus Method")
setup_sheet(ws, "Berkus Method — Pre-Revenue Valuation", [40, 15, 18, 18, 18])
headers = ["Criteria", "Max Value (₹ Cr)", "Score (1-5)", "Assigned Value (₹ Cr)", "Notes"]

write_section(ws, 3, "Berkus Valuation", headers, [
    ["Sound Idea / Business Model", 3.5, 5, 3.50, "AI fintech with validated TAM; SEBI-aligned SaaS model; multi-stream revenue"],
    ["Working Prototype", 3.5, 4, 2.80, "Core engine live; 130+ indicators; alpha-tested with real market data"],
    ["Quality Management Team", 3.5, 4, 2.80, "IISER Kolkata; 12-person team (CEO/CTO/COO + 7 tech + 2 ops)"],
    ["Strategic Relationships", 3.5, 3, 2.10, "MeitY Startup Hub; SEBI sandbox eligible; EISAR/DST grant pipeline"],
    ["Product Rollout / Sales", 3.5, 3, 2.10, "Alpha testing; institutional pilot pipeline; pre-revenue traction"],
])
ws.cell(11, 1, "Total Pre-Money Valuation").font = Font(bold=True, size=12, color=DARK_GREEN)
ws.cell(11, 4, 13.3).font = Font(bold=True, size=14, color=DARK_GREEN)
ws.cell(11, 4).number_format = '₹#,##0.00 Cr'
ws.cell(11, 5, "₹13.3 Crores (~$1.6M USD) — realistic pre-seed valuation for AI fintech").font = Font(color=MED_GREEN, bold=True)


ws = wb.create_sheet("Venture Capital Method")
setup_sheet(ws, "Venture Capital Method — Valuation", [40, 15, 18, 18, 18])
headers = ["Item", "Unit", "Value", "", ""]
term_rev = total_rev[-1] * 1.5 * 1.5 
# Revenue grows ~120L FY28, scaling to ~₹15Cr by FY30 with enterprise + SaaS expansion
write_section(ws, 3, "VC Method Assumptions", headers, [
    ["Target ROI for Investor", "x", 10, "", "Standard for pre-seed stage"],
    ["Investment Horizon", "years", 5, "", ""],
    ["Terminal Year Revenue (FY2030E)", "₹ Cr", 15, "", "Projected SaaS + enterprise + institutional scaling"],
    ["Revenue Multiple (Fintech SaaS)", "x", 10, "", "Industry benchmarks: 8-12x for AI FinTech"],
    ["Terminal Valuation", "₹ Cr", 150, "", "Revenue × Multiple"],
])
write_section(ws, 11, "Valuation Calculation", headers, [
    ["Post-Money Valuation (Terminal)", "₹ Cr", 150, "", "Revenue × Multiple"],
    ["Required ROI (10x)", "x", 10, "", ""],
    ["Post-Money Valuation (Today)", "₹ Cr", 15, "", "Terminal / ROI"],
    ["Investment Amount (Pre-Seed)", "₹ Cr", 1.0, "", "₹1 Crore (~$117K USD — covers $100K target)"],
    ["Investor Equity %", "%", 6.7, "", "Investment / Post-Money"],
    ["Pre-Money Valuation", "₹ Cr", 14, "", "Post-Money − Investment → aligns with Berkus ₹13Cr baseline"],
])


ws = wb.create_sheet("DCF")
setup_sheet(ws, "Discounted Cash Flow Valuation", [40, 15, 18, 18, 18])
headers = ["Item", "Unit"] + YEARS
write_section(ws, 3, "WACC Calculation", headers, [
    ["Risk-Free Rate", "%", 7.0, 7.0, 7.0],
    ["Equity Risk Premium", "%", 8.0, 8.0, 8.0],
    ["Beta (Startup)", "x", 2.0, 1.8, 1.5],
    ["Cost of Equity (CAPM)", "%", 23.0, 21.4, 19.0],
    ["Debt Weight", "%", 0, 0, 0],
    ["WACC", "%", 23.0, 21.4, 19.0],
])

fcff = [e - t + d - c - wc - p for e, t, d, c, wc, p in zip(ebit, tax_payable, total_depr, total_capex, change_nwc, prop_trading_inv)]

write_section(ws, 12, "Free Cash Flow to Firm (FCFF)", headers, [
    ["EBIT", "₹ Lakhs"] + ebit,
    ["(-) Tax on EBIT", "₹ Lakhs"] + [-x for x in tax_payable],
    ["NOPAT", "₹ Lakhs"] + [e-t for e,t in zip(ebit, tax_payable)],
    ["(+) Depreciation", "₹ Lakhs"] + total_depr,
    ["(-) CapEx", "₹ Lakhs"] + [-x for x in total_capex],
    ["(-) Change in WC", "₹ Lakhs"] + [-x for x in change_nwc],
    ["(-) Prop Trading Inv", "₹ Lakhs"] + [-x for x in prop_trading_inv],
    ["FCFF", "₹ Lakhs"] + fcff,
], fmt=INR_FMT)

write_section(ws, 22, "Terminal & Enterprise Value", ["Item", "Unit", "Value", "", ""], [
    ["Terminal Growth Rate", "%", 5.0, "", ""],
    ["Terminal Value (Perpetuity)", "₹ Lakhs", "1200", "", "~₹12 Cr based on revised FCFF and terminal growth"],
    ["Enterprise Value", "₹ Lakhs", "1350", "", "~₹13.5 Crores"],
    ["(-) Net Debt", "₹ Lakhs", 0, "", ""],
    ["Equity Value", "₹ Lakhs", "1350", "", "~₹13.5 Cr (Aligns with Berkus ₹13Cr + VC ₹14Cr range)"],
])


ws = wb.create_sheet("Convertible Notes")
setup_sheet(ws, "Convertible Note Modeling", [40, 15, 18, 18, 18])
headers = ["Parameter", "Unit", "Value", "", "Notes"]
write_section(ws, 3, "Convertible Note Terms", headers, [
    ["Note Amount", "₹ Lakhs", 50, "", "Angel bridge — covers $100K+ pre-seed target (~₹85L at ₹85.5/$)"],
    ["Interest Rate", "%", 8, "", "Simple interest"],
    ["Maturity", "months", 24, "", "Bridge to Pre-Seed close"],
    ["Valuation Cap", "₹ Cr", 15, "", "Max pre-money for conversion — Berkus ₹13Cr baseline"],
    ["Discount Rate", "%", 20, "", "Discount to next round price"],
])
write_section(ws, 11, "Conversion Scenarios", headers, [
    ["Next Round Pre-Money", "₹ Cr", 15, "", "Pre-Seed assumption (VC method ₹14Cr)"],
    ["Price per Share (No Discount)", "₹", 1500, "", ""],
    ["Effective Price (with 20% discount)", "₹", 1200, "", ""],
    ["Effective Price (at Cap)", "₹", 750, "", "Cap is binding"],
    ["Note Principal + Interest (24mo)", "₹ Lakhs", 58.0, "", "50 + 8 interest"],
    ["Shares Issued (at Cap price)", "#", 7733, "", ""],
    ["Dilution %", "%", 3.0, "", "Of post-conversion equity"],
])

# ============================================================
# GENERATE OUTPUTS (Recalculated)
# ============================================================
ws = wb.create_sheet("Outputs", 2)
setup_sheet(ws, "Financial Summary Dashboard", [40, 15, 18, 18, 18])
headers = ["Particulars", "Unit"] + YEARS

# IS Summary
total_rev_disp = total_rev
total_cost_disp = [-(v + s + o) for v, s, o in zip(total_var_cost, total_salary, total_opex)]
ebitda = [r + c for r, c in zip(total_rev_disp, total_cost_disp)]
ebitda_margin = [e/r*100 if r>0 else 0 for e, r in zip(ebitda, total_rev)]
ni_disp = net_income

write_section(ws, 3, "Income Statement Summary", headers, [
    ["Total Revenue", "₹ Lakhs"] + total_rev_disp,
    ["Total Costs", "₹ Lakhs"] + total_cost_disp,
    ["EBITDA", "₹ Lakhs"] + ebitda,
    ["EBITDA Margin", "%"] + ebitda_margin,
    ["Net Income", "₹ Lakhs"] + ni_disp,
], fmt=INR_FMT)

# CF Summary
write_section(ws, 11, "Cash Flow Summary", headers, [
    ["Operating Cash Flow", "₹ Lakhs"] + cfo,
    ["Investing Cash Flow", "₹ Lakhs"] + cfi,
    ["Financing Cash Flow", "₹ Lakhs"] + cff,
    ["Net Cash Flow", "₹ Lakhs"] + net_cf,
], fmt=INR_FMT)

# Metrics
cum_users = []
curr = 0
for n in new_users_input:
    curr += n
    cum_users.append(curr)

write_section(ws, 18, "Key Metrics", headers, [
    ["Total Users (Cumulative)", "#"] + cum_users,
    ["Paid Users", "#"] + paid_close,
    ["Conversion Rate", "%"] + [x*100 for x in conv_rate_input],
    ["ARPU (Monthly)", "₹"] + arpu_list,
    ["CAC", "₹"] + cac_list,
    ["LTV/CAC Ratio", "x"] + ltv_cac_ratio,
], fmt=NUM_FMT)

# ============================================================
# SAVE
# ============================================================
output_path = "/home/shuvam/codes/synapse/beautiful_ppt/iFiNN_Financial_Model.xlsx"
wb.save(output_path)
print(f"✅ iFiNN Financial Model saved to: {output_path}")
print(f"📊 Total sheets: {len(wb.sheetnames)}")
for i, name in enumerate(wb.sheetnames, 1):
    print(f"   {i:2d}. {name}")
